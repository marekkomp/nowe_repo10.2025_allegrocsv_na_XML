# scripts/convert_Morele.py
import os
from convert import convert_file, INPUT_DIR, OUTPUT_DIR, _desc_to_html, _is_available, _looks_like_json
import xml.etree.ElementTree as ET
import openpyxl

def convert_file_morele(in_path, out_path):
    # Wczytujemy tak samo jak w głównym konwerterze
    wb = openpyxl.load_workbook(in_path, read_only=True, data_only=True)
    ws = wb["Szablon"] if "Szablon" in wb.sheetnames else wb.worksheets[0]
    headers = [str(c).strip() if c else "" for c in next(ws.iter_rows(min_row=4, max_row=4, values_only=True))]
    rows = list(ws.iter_rows(min_row=5, values_only=True))

    # Szukamy potrzebnych kolumn
    i_id   = headers.index("ID oferty")
    i_tyt  = headers.index("Tytuł oferty")
    i_cena = headers.index("Cena PL")
    i_url  = headers.index("Link do oferty")
    i_stat = headers.index("Status oferty")
    i_qty  = headers.index("Liczba sztuk")
    i_cat  = headers.index("Kategoria główna")
    i_desc = headers.index("Opis oferty")

    root = ET.Element("offers")

    for row in rows:
        id_offer = str(row[i_id]).strip() if row[i_id] else ""
        title    = str(row[i_tyt]).strip() if row[i_tyt] else ""
        price    = str(row[i_cena]).strip() if row[i_cena] else ""
        url      = str(row[i_url]).strip() if row[i_url] else ""
        status   = str(row[i_stat]).strip().lower() if row[i_stat] else ""
        qty      = float(row[i_qty]) if row[i_qty] else 0
        cat      = str(row[i_cat]).strip() if row[i_cat] else ""
        desc_raw = str(row[i_desc]).strip() if row[i_desc] else ""

        if not id_offer or not title:
            continue

        # 🔹 Zmieniona logika dostępności
        available = (status == "aktywna") and (qty >= 5)

        avail_val = "1" if available else "99"
        stock     = "1" if available else "0"
        basket    = "1" if available else "0"

        o = ET.SubElement(root, "o", {
            "id": id_offer,
            "url": url,
            "price": price,
            "avail": avail_val,
            "stock": stock,
            "basket": basket,
        })

        ET.SubElement(o, "cat").text = cat or ""
        ET.SubElement(o, "name").text = title

        # 🔹 Bez desc_json
        if desc_raw:
            desc_el = ET.SubElement(o, "desc")
            desc_el.text = _desc_to_html(desc_raw, strict=True)

    ET.indent(root, space="  ")
    ET.ElementTree(root).write(out_path, encoding="utf-8", xml_declaration=True)
    print(f"[Morele OK] Zapisano: {out_path}")

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    for name in os.listdir(INPUT_DIR):
        if name.lower().endswith((".xlsx", ".xlsm", ".xls")):
            src = os.path.join(INPUT_DIR, name)
            dst = os.path.join(OUTPUT_DIR, "morele.xml")
            print(f"[Morele] {src} -> {dst}")
            convert_file_morele(src, dst)
            break

if __name__ == "__main__":
    main()
