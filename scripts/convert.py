# scripts/convert.py
import json
import html
import re as _re
import os
import re
import xml.etree.ElementTree as ET
import openpyxl

INPUT_DIR = "input"
OUTPUT_DIR = "output"
DESC_STRICT = True  # bez „upiększania”; składamy JSON->HTML + lekka sanizacja


# Pola wymagane do znalezienia danych
REQ_HEADERS = ["Tytuł oferty", "Cena PL", "Link do oferty", "Status oferty", "Liczba sztuk", "ID oferty"]

# Mapa kolumn Excela do atrybutów XML (puste są automatycznie pomijane)
ATTR_MAP = {
    # Identyfikacja
    "Producent": "Producent",
    "Kod producenta": "kod_producenta",
    "Model": "model_1",
    "ID produktu (EAN/UPC/ISBN/ISSN/ID produktu Allegro)": "ean",
    "Sygnatura/SKU Sprzedającego": "sku",

    # CPU / RAM
    "Model procesora": "model_procesora",
    "Generacja procesora": "generacja_procesora",
    "Taktowanie bazowe procesora [GHz]": "taktowanie_bazowe_procesora",
    "Taktowanie maksymalne procesora [GHz]": "taktowanie_maksymalne_procesora",
    "Liczba rdzeni procesora": "liczba_rdzeni_procesora",
    "Liczba wątków procesora": "liczba_watkow_procesora",
    "Pamięć podręczna procesora [MB]": "pamiec_podreczna_procesora_mb",
    "Typ pamięci RAM": "typ_pamieci_ram",
    "Wielkość pamięci RAM": "wielkosc_pamieci_ram",
    "Taktowanie szyny pamięci RAM [MHz]": "taktowanie_pamieci_ram_mhz",
    "Maksymalna pojemność pamięci RAM [GB]": "max_pamiec_ram_gb",

    # Dysk / magazyn
    "Typ dysku twardego": "typ_dysku",
    "Pojemność dysku [GB]": "pojemnosc_dysku_gb",
    "Format dysku": "format_dysku",
    "Interfejs dysku": "interfejs_dysku",
    "Prędkość obrotowa dysku HDD": "predkosc_hdd_rpm",

    # Grafika
    "Producent karty graficznej": "producent_karty_graficznej",
    "Chipset karty graficznej": "chipset_karty_graficznej",
    "Pamięć karty graficznej": "pamiec_karty_graficznej",
    "Złącza karty graficznej": "zlacza_karty_graficznej",
    "Rodzaj karty graficznej": "rodzaj_karty_graficznej",

    # Ekran
    "Przekątna ekranu (cale) [\"]": "przekatna_ekranu",
    "Rozdzielczość natywna [px]": "rozdzielczosc_ekranu",
    "Typ matrycy": "typ_matrycy",
    "Powłoka matrycy": "powloka_matrycy",
    "Proporcje obrazu": "proporcje_obrazu",
    "Częstotliwość odświeżania [Hz]": "odswiezanie_hz",

    # Wejścia/wyjścia i łączność
    "Złącza": "zlacza_zewnetrzne",
    "Komunikacja": "komunikacja",
    "Standard HDMI": "standard_hdmi",

    # System / klawiatura / zasilanie
    "System operacyjny": "system_operacyjny",
    "Wersja systemu operacyjnego": "wersja_systemu_operacyjnego",
    "Typ klawiatury": "typ_klawiatury",
    "Układ klawiatury": "uklad_klawiatury",
    "Ładowarka w komplecie": "ladowarka_w_komplecie",

    # Stan / gwarancja
    "Stan": "stan_urządzenia",
    "Stan opakowania": "stan_opakowania",
    "Dołączone oprogramowanie": "dolaczone_oprogramowanie",
    "Informacje o gwarancjach (opcjonalne)": "gwarancja_info",
}


def _clean_headers(cells):
    hdr = [("" if v is None else str(v).strip()) for v in cells]
    while hdr and hdr[-1] == "":
        hdr.pop()
    return hdr

def _read_headers_stream(ws, max_col=500):
    row = next(ws.iter_rows(min_row=4, max_row=4, max_col=max_col, values_only=True))
    return _clean_headers(row)

def _read_headers_full(ws):
    max_col = ws.max_column if ws.max_column and ws.max_column > 0 else 500
    cells = [ws.cell(row=4, column=c).value for c in range(1, max_col + 1)]
    return _clean_headers(cells)

def _idx(headers, name, default=-1):
    return headers.index(name) if name in headers else default

def _as_str(val):
    return "" if val is None else str(val).strip()

def _parse_images(raw):
    if not raw:
        return []
    parts = [u.strip() for u in str(raw).split("|") if u.strip()]
    urls = [u for u in parts if re.match(r"^https?://", u)]
    return urls
def _sanitize_html_basic(s: str) -> str:
    """Lekka sanizacja – usuń <script> i <iframe>; resztę zostaw."""
    if not s:
        return s
    s = _re.sub(r"(?is)<script.*?</script>", "", s)
    s = _re.sub(r"(?is)<iframe.*?</iframe>", "", s)
    return s

def _looks_like_json(s: str) -> bool:
    if not s:
        return False
    s = s.lstrip()
    return s.startswith("{") or s.startswith("[")

def _desc_to_html(desc_raw: str, strict: bool = True) -> str:
    """
    Zamienia JSON `{"sections":[{"items":[...]}]}` na HTML.
    Jeżeli to nie JSON – zwraca lekko zsanitowany oryginał.
    """
    if not desc_raw:
        return ""
    s = desc_raw.strip()
    if not _looks_like_json(s):
        return _sanitize_html_basic(s)

    try:
        data = json.loads(s)
    except Exception:
        # Niepoprawny JSON – zwróć oryginał (po sanizacji)
        return _sanitize_html_basic(desc_raw)

    try:
        raw_sections = data.get("sections", [])
    except AttributeError:
        raw_sections = data if isinstance(data, list) else []

    sections_html = []
    for sec in raw_sections:
        items = sec.get("items", []) if isinstance(sec, dict) else []
        parts = []
        for it in items:
            t = (it.get("type") or "").upper()
            if t == "TEXT":
                parts.append(it.get("content", ""))  # content to już HTML
            elif t == "IMAGE":
                url = (it.get("url") or "").strip()
                if url:
                    parts.append(f'<p><img src="{html.escape(url, quote=True)}" loading="lazy" alt=""></p>')
        if parts:
            sections_html.append("\n".join(parts))

    html_out = "\n\n".join(sections_html)

    if not strict:
        html_out = _re.sub(r"<h1>[_\-–—\s]{3,}</h1>", "<hr>", html_out, flags=_re.IGNORECASE)
        html_out = _re.sub(r"(?i)<h1>", "<h2>", html_out)
        html_out = _re.sub(r"(?i)</h1>", "</h2>", html_out)

    return _sanitize_html_basic(html_out)

def _is_available(status, qty):
    try:
        q = int(float(str(qty).replace(",", ".").strip())) if str(qty).strip() != "" else 0
    except:
        q = 0
    return (str(status).strip().lower() == "aktywna") and (q > 0)

def _ensure_required(headers):
    return [h for h in REQ_HEADERS if h not in headers]

def convert_file(in_path, out_path):
    # 1) próba streaming (read_only)
    wb = openpyxl.load_workbook(in_path, read_only=True, data_only=True)
    ws = wb["Szablon"] if "Szablon" in wb.sheetnames else wb.worksheets[0]

    headers = _read_headers_stream(ws, max_col=500)
    print(f"[INFO] Arkusz: {ws.title}")
    print(f"[DEBUG] Nagłówków (stream): {len(headers)}")
    print(f"[DEBUG] Podgląd (stream): {headers[:30]}")
    missing = _ensure_required(headers)

    # 2) jeśli brakuje kolumn — tryb pełny
    if missing:
        print(f"[WARN] Brak w stream: {missing} → pełny tryb")
        wb.close()
        wb = openpyxl.load_workbook(in_path, data_only=True)  # bez read_only
        ws = wb["Szablon"] if "Szablon" in wb.sheetnames else wb.worksheets[0]
        headers = _read_headers_full(ws)
        print(f"[DEBUG] Nagłówków (full): {len(headers)}")
        print(f"[DEBUG] Podgląd (full): {headers[:30]}")
        missing = _ensure_required(headers)

    if missing:
        print(f"[ERROR] Brak wymaganych kolumn nawet w trybie pełnym: {missing}")
        ET.ElementTree(ET.Element("offers")).write(out_path, encoding="utf-8", xml_declaration=True)
        return

    # Indeksy
    i_id     = _idx(headers, "ID oferty")
    i_title  = _idx(headers, "Tytuł oferty")
    i_price  = _idx(headers, "Cena PL")
    i_url    = _idx(headers, "Link do oferty")
    i_stat   = _idx(headers, "Status oferty")
    i_qty    = _idx(headers, "Liczba sztuk")
    i_cat    = _idx(headers, "Kategoria główna")
    i_sub    = _idx(headers, "Podkategoria")
    i_imgs   = _idx(headers, "Zdjęcia")
    i_desc   = _idx(headers, "Opis oferty")  # [NOWE]

    # Zrzut wierszy (data_only=True)
    rows = list(ws.iter_rows(min_row=5, values_only=True))
    if not any(any(_as_str(c) for c in r) for r in rows):
        print("[WARN] Arkusz wygląda na formułowy (data_only puste). Odczyt z data_only=False.")
        wb.close()
        wb = openpyxl.load_workbook(in_path, data_only=False)
        ws = wb["Szablon"] if "Szablon" in wb.sheetnames else wb.worksheets[0]
        rows = list(ws.iter_rows(min_row=5, values_only=True))

    root = ET.Element("offers")

    # Główna pętla po ofertach
    offers_count = 0
    for row in rows:
        # bezpieczeństwo
        if max(i_id, i_title, i_price, i_url, i_stat, i_qty) >= len(row):
            continue

        id_offer = _as_str(row[i_id])
        title    = _as_str(row[i_title])
        price    = _as_str(row[i_price])
        url      = _as_str(row[i_url])
        status   = _as_str(row[i_stat])
        qty      = row[i_qty] if i_qty < len(row) else ""
        cat      = _as_str(row[i_cat]) if i_cat < len(row) else ""
        subcat   = _as_str(row[i_sub]) if i_sub < len(row) else ""
        imgs_raw = row[i_imgs] if i_imgs < len(row) else ""
        desc_raw = _as_str(row[i_desc]) if (i_desc != -1 and i_desc < len(row)) else ""  # [NOWE]

        # pomijamy bez ID i bez tytułu  [NOWE]
        if not id_offer or not title:
            continue

        available = _is_available(status, qty)
        avail_val = "1" if available else "99"
        basket    = "1" if available else "0"
        stock     = "1" if available else "0"

        # element <o ...>
        o = ET.SubElement(
            root,
            "o",
            {
                "id": id_offer,
                "url": url,
                "price": price,
                "avail": avail_val,
                "stock": stock,
                "basket": basket,
            },
        )

        # <cat> tylko Kategoria główna  [ZMIANA]
        ET.SubElement(o, "cat").text = cat or ""

        # <name>
        ET.SubElement(o, "name").text = title

        # <desc_json> (jeśli surowy JSON) + <desc> (HTML)
if desc_raw:
    if _looks_like_json(desc_raw):
        desc_json_el = ET.SubElement(o, "desc_json")
        desc_json_el.text = desc_raw  # surowy JSON bez zmian

    desc_el = ET.SubElement(o, "desc")
    desc_el.text = _desc_to_html(desc_raw, strict=DESC_STRICT)


        # <imgs>
        imgs = _parse_images(imgs_raw)
        imgs_el = ET.SubElement(o, "imgs")
        if imgs:
            ET.SubElement(imgs_el, "main", {"url": imgs[0]})
            for u in imgs[1:]:
                ET.SubElement(imgs_el, "i", {"url": u})

        # <attrs> – tylko wypełnione pola z mapy
        attrs_el = ET.SubElement(o, "attrs")
        for col, attr_name in ATTR_MAP.items():
            if col in headers:
                idx = headers.index(col)
                if idx < len(row):
                    val = _as_str(row[idx])
                    if val:
                        ET.SubElement(attrs_el, "a", {"name": attr_name}).text = val

        offers_count += 1

    ET.indent(root, space="  ")
    ET.ElementTree(root).write(out_path, encoding="utf-8", xml_declaration=True)
    print(f"[OK] Zapisano: {out_path} | ofert: {offers_count}")

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    any_processed = False
    for name in os.listdir(INPUT_DIR):
        if name.lower().endswith((".xlsm", ".xlsx", ".xls")):
            src = os.path.join(INPUT_DIR, name)
            dst = os.path.join(OUTPUT_DIR, os.path.splitext(name)[0] + ".xml")
            print(f"[RUN] {src} -> {dst}")
            convert_file(src, dst)
            any_processed = True
    if not any_processed:
        print("[INFO] Brak plików wejściowych w /input")

if __name__ == "__main__":
    main()
